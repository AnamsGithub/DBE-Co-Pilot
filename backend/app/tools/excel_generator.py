from pathlib import Path
from copy import deepcopy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.models.schemas import DBEResumeData, PointerImprovement

TEMPLATE_PATH = Path("/Users/apple/Downloads/CV_Format(Junior) (1).xlsx")
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# Exact colors from template inspection
GREY_HEADER   = PatternFill("solid", fgColor="FFBEBEBE")   # section headers
GREY_SUBROW   = PatternFill("solid", fgColor="FFD9D9D9")   # company/position rows
BLACK_FILL    = PatternFill("solid", fgColor="FF1F2937")   # tagline (dark)
NO_FILL       = PatternFill("none")

# Column widths from template (character units)
COL_WIDTHS = {
    "A": 27.6, "B": 21.6, "C": 7.2, "D": 16.6,
    "E": 20.2, "F": 18.6, "G": 13.6, "H": 11.6,
}

ROW_H_NAME   = 27.0
ROW_H_INFO   = 18.0
ROW_H_NORMAL = 22.05
ROW_H_SPACE  = 5.0

ARIAL = "Arial"


def _f(size=12, bold=False, color="FF000000", italic=False):
    return Font(name=ARIAL, size=size, bold=bold, color=color, italic=italic)


def _a(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(left=True, right=True, top=False, bottom=False):
    thin = Side(style="thin")
    none = Side(style=None)
    return Border(
        left=thin if left else none,
        right=thin if right else none,
        top=thin if top else none,
        bottom=thin if bottom else none,
    )


def _set(ws, coord, value="", bold=False, h="left", italic=False,
         fill=None, font_color="FF000000", border=None, size=12):
    c = ws[coord]
    c.value = value
    c.font = _f(size=size, bold=bold, color=font_color, italic=italic)
    c.alignment = _a(h=h)
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border


def _merge(ws, r1, c1, r2, c2, value="", bold=False, h="left", italic=False,
           fill=None, font_color="FF000000", border=None, size=12):
    from openpyxl.utils import get_column_letter as gcl
    start = f"{gcl(c1)}{r1}"
    end   = f"{gcl(c2)}{r2}"
    ws.merge_cells(f"{start}:{end}")
    c = ws[start]
    c.value = value
    c.font = _f(size=size, bold=bold, color=font_color, italic=italic)
    c.alignment = _a(h=h)
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border


def _row_h(ws, row, height):
    ws.row_dimensions[row].height = height


def _apply_improved(bullets, improvements, section_key):
    imp_map = {i.original: i.improved for i in improvements
               if section_key.lower() in i.section.lower()}
    return [imp_map.get(b, b) for b in bullets]


# ─────────────────────────────────────────────────────────────────────────────

def generate_cv_excel(data: DBEResumeData, improvements: list[PointerImprovement]) -> str:
    # Load template to copy logo image and base styles
    tmpl = load_workbook(str(TEMPLATE_PATH))
    ws_tmpl = tmpl.active

    wb = Workbook()
    ws = wb.active
    ws.title = "CV"

    # Set column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Copy logo image from template
    for img in ws_tmpl._images:
        try:
            from copy import deepcopy
            new_img = deepcopy(img)
            ws.add_image(new_img)
        except Exception:
            pass

    row = 1

    # ── Row 1: Name + Department ──────────────────────────────────────────────
    _row_h(ws, row, ROW_H_NAME)
    _merge(ws, row, 1, row, 4, value=data.name, bold=True, size=24)   # A1:D1
    _merge(ws, row, 5, row+1, 8,                                       # E1:H2
           value="Department of Business Economics", bold=True, size=16, h="right")
    row += 1

    # ── Row 2: Gender / Age / Languages ───────────────────────────────────────
    _row_h(ws, row, ROW_H_INFO)
    _merge(ws, row, 1, row, 4,                                          # A2:D2
           value=f"{data.gender}, {data.age} years | {data.languages}", bold=True)
    row += 1

    # ── Row 3: Contact ────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_INFO)
    _merge(ws, row, 1, row, 4, value=f"{data.email} | {data.phone}")   # A3:D3
    _merge(ws, row, 6, row, 8, value="placement@dbe-du.org", h="right") # F3:H3
    row += 1

    # ── Row 4: Tagline ────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _merge(ws, row, 1, row, 8, value=data.tagline,                      # A4:H4
           bold=True, h="center", fill=BLACK_FILL, font_color="FFFFFFFF")
    row += 1

    # ── Row 5: Educational Qualifications header ──────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _merge(ws, row, 1, row, 8, value="Educational Qualifications",      # A5:H5
           bold=True, fill=GREY_HEADER)
    row += 1

    # ── Education rows ────────────────────────────────────────────────────────
    for edu in data.education:
        _row_h(ws, row, ROW_H_NORMAL)
        _merge(ws, row, 1, row, 2, value=edu.get("degree", ""))         # A:B
        _set(ws,  f"C{row}", value=edu.get("year", ""), h="center")     # C
        _merge(ws, row, 4, row, 6, value=edu.get("institution", ""), h="center")  # D:F
        _merge(ws, row, 7, row, 8, value=edu.get("grade", ""), h="center")        # G:H
        row += 1

    # ── Spacing ───────────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Professional Experience ───────────────────────────────────────────────
    if data.experience:
        durations = [e.get("duration", "") for e in data.experience if e.get("duration")]
        total_label = durations[0] if len(durations) == 1 else f"{sum(int(d.split()[0]) for d in durations if d.split()[0].isdigit()):02d} Months"
        _row_h(ws, row, ROW_H_NORMAL)
        _merge(ws, row, 1, row, 7, value="Professional Experience",     # A:G
               bold=True, fill=GREY_HEADER)
        _set(ws, f"H{row}", value=total_label, bold=True, h="center", fill=GREY_HEADER)
        row += 1

        for exp in data.experience:
            bullets = _apply_improved(exp.get("bullets", []), improvements, exp.get("company", ""))
            row = _entry_block(ws, row, exp.get("company", ""), exp.get("role", ""), exp.get("dates", ""), bullets)

        _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Internship ────────────────────────────────────────────────────────────
    if data.internships:
        durations = [i.get("duration", "") for i in data.internships if i.get("duration")]
        total_label = durations[0] if len(durations) == 1 else f"{sum(int(d.split()[0]) for d in durations if d.split()[0].isdigit()):02d} Weeks"
        _row_h(ws, row, ROW_H_NORMAL)
        _merge(ws, row, 1, row, 7, value="Internship", bold=True, fill=GREY_HEADER)
        _set(ws, f"H{row}", value=total_label, bold=True, h="center", fill=GREY_HEADER)
        row += 1

        for intern in data.internships:
            bullets = _apply_improved(intern.get("bullets", []), improvements, intern.get("company", ""))
            row = _entry_block(ws, row, intern.get("company", ""), intern.get("role", ""), intern.get("dates", ""), bullets)

        _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Key Projects ──────────────────────────────────────────────────────────
    if data.projects:
        _row_h(ws, row, ROW_H_NORMAL)
        _merge(ws, row, 1, row, 8, value="Key Projects", bold=True, fill=GREY_HEADER)
        row += 1

        for proj in data.projects:
            bullets = _apply_improved(proj.get("bullets", []), improvements, proj.get("name", ""))
            n = len(bullets)
            if n == 0:
                continue
            # Merge project name cell across all its bullet rows
            _merge(ws, row, 1, row + n - 1, 1, value=proj.get("name", ""), fill=GREY_SUBROW)
            # Merge domain cell across all bullet rows
            _merge(ws, row, 8, row + n - 1, 8, value=proj.get("skill_tag", ""),
                   fill=GREY_SUBROW, h="center")
            for b in bullets:
                _row_h(ws, row, ROW_H_NORMAL)
                _merge(ws, row, 2, row, 7, value=f"·  {b}",          # B:G
                       border=_border(left=True, right=True))
                row += 1

        _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Positions of Responsibilities ─────────────────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _merge(ws, row, 1, row, 8, value="Positions of Responsibilities",
           bold=True, fill=GREY_HEADER)
    row += 1

    for pos in data.positions:
        bullets = _apply_improved(pos.get("bullets", []), improvements, pos.get("role", ""))
        n = len(bullets)
        if n == 0:
            continue
        _merge(ws, row, 1, row + n - 1, 1, value=pos.get("role", ""), fill=GREY_SUBROW)
        _merge(ws, row, 7, row + n - 1, 8, value=pos.get("dates", ""),
               fill=NO_FILL, h="center")
        for b in bullets:
            _row_h(ws, row, ROW_H_NORMAL)
            _merge(ws, row, 2, row, 6, value=f"·  {b}",             # B:F
                   border=_border(left=True, right=True))
            row += 1

    _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Achievements ──────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _merge(ws, row, 1, row, 8, value="Achievements", bold=True, fill=GREY_HEADER)
    row += 1

    for label, items in [
        ("Academic Achievements", data.academic_achievements),
        ("Competitions/ Events",  data.competition_achievements),
        ("Certifications",        data.certifications),
    ]:
        if not items:
            continue
        n = len(items)
        _merge(ws, row, 1, row + n - 1, 1, value=label, fill=GREY_SUBROW)
        for item in items:
            _row_h(ws, row, ROW_H_NORMAL)
            _merge(ws, row, 2, row, 8, value=f"·  {item}",          # B:H
                   border=_border(left=True, right=True))
            row += 1

    _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Interests ─────────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _set(ws, f"A{row}", value="Interests", bold=True, fill=GREY_HEADER)
    _merge(ws, row, 2, row, 8, value=" " + " | ".join(data.interests))   # B:H
    row += 1

    _row_h(ws, row, ROW_H_SPACE); row += 1

    # ── Skills ────────────────────────────────────────────────────────────────
    _row_h(ws, row, ROW_H_NORMAL)
    _set(ws, f"A{row}", value="Skills", bold=True, fill=GREY_HEADER)
    _merge(ws, row, 2, row, 8, value=" " + " | ".join(data.skills))      # B:H

    # Save
    safe_name = data.name.strip().replace(" ", "_")
    out_name = f"{safe_name}_CV_corrected.xlsx"
    out_path = OUTPUT_DIR / out_name
    wb.save(str(out_path))
    return out_name


def _entry_block(ws, row, company, role, dates, bullets):
    """Company header row + bullet rows matching DBE-DU template structure."""
    _row_h(ws, row, ROW_H_NORMAL)
    _set(ws, f"A{row}", value=company, bold=True, fill=GREY_SUBROW)     # A alone
    _merge(ws, row, 2, row, 6, value=role, bold=True,                   # B:F
           fill=GREY_SUBROW, h="center")
    _merge(ws, row, 7, row, 8, value=dates, bold=True,                  # G:H
           fill=GREY_SUBROW, h="center")
    row += 1

    for b in bullets:
        _row_h(ws, row, ROW_H_NORMAL)
        _merge(ws, row, 2, row, 8, value=f"·  {b}",               # B:H
               border=_border(left=True, right=True))
        row += 1

    return row
